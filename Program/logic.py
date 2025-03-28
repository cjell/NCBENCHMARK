# This is the logic.py file. It contains the main logic for the anomaly detection program.

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Define fill colors.
green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
red = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
blue = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Runs the anomaly detection algorithm on the provided CSV file.
# Parameters:
#   csv_path: Path to the input CSV file.
#   threshold: Z-score threshold for anomaly detection.
#   output_path: Path to the output Excel file.
#   split_by: If set to "municipality" or "category", the pivoted (transformed) data
#             will be split into separate sheets by that column. If None, all transformed
#             data is written to a single sheet ("Transformed Data").
#   target_year: The year to analyze (e.g., 2024).
def run_anomaly_detection(csv_path, threshold, output_path, split_by, target_year):
    
    # Reads CSV with two header rows (codes + actual names)
    df = pd.read_csv(csv_path, header=[0, 1], thousands=',')
    
    # Identify important columns used for indexing
    year_col = 'a_fiscal_year'
    muni_col = 'a_jurisdiction'
    category_col = 'a_service'
    
    # Creates a dictionary mapping from metric code (key) to actual metric name (value)
    metric_names = {}
    cols = df.columns.tolist()  # list of tuples (code, actual_name)
    for i, col in enumerate(cols):
        if i >= 3:  # Since the first three columns are not metrics
            metric_code = col[0]
            metric_name = col[1] if pd.notna(col[1]) else ""
            metric_names[metric_code] = metric_name

    # Flatten header: keep only the first row (the codes)
    df.columns = [col[0] for col in cols]
    metric_cols = df.columns[3:]

    # Create a filtered DataFrame for detection: ignore rows missing a category.
    df_detection = df[df[category_col].notna()].copy()

    # Detection Loop (only for the target_year, using df_detection)
    anomalies = []      # Outliers (to be highlighted red)
    missing_data = []   # Missing value (highlight blue)
    non_numeric = []    # Non-numeric values (highlight yellow)

    for metric in metric_cols: #iterating over columns
        for idx, row in df_detection.iterrows(): # Iterating over rows
            raw_year = row[year_col]
            raw_muni = row[muni_col]
            raw_category = row[category_col]
            try:
                current_year = float(raw_year)
            except ValueError:
                continue # Skip rows with non-numeric year
            if current_year != target_year:
                continue

            val = row[metric]
            # Excel-like coordinates (assume header occupies row 1)
            excel_row = idx + 2
            excel_col = df_detection.columns.get_loc(metric) + 1

            # Check for non-numeric value
            if not pd.isna(val):
                try:
                    current_val = float(val)
                except ValueError:
                    non_numeric.append((excel_row, excel_col, val, raw_year, raw_muni, raw_category, metric))
                    continue
            else:
                current_val = None

            # Gather historical data (for years < target_year) for same municipality and category
            hist_rows = df_detection.loc[
                (df_detection[year_col] < current_year) &
                (df_detection[muni_col] == raw_muni) &
                (df_detection[category_col] == raw_category),
                [year_col, category_col, metric]
            ].dropna()
            try:
                historical_data = hist_rows[metric].astype(float).values
            except ValueError:
                historical_data = np.array([])

            hist_comment = "\n".join(
                f"{int(r[year_col])}: {r[metric]}"
                for _, r in hist_rows.iterrows()
            )

            # If value is missing but history exists, flag as missing.
            if pd.isna(val) and len(historical_data) > 0:
                missing_data.append((excel_row, excel_col, raw_year, raw_muni, raw_category, metric, hist_comment))
                continue

            if len(historical_data) < 2 or current_val is None:
                continue

            mean = (historical_data).mean()
            std_dev = (historical_data).std()
            z_score = (current_val - mean) / std_dev if std_dev != 0 else 0

            if abs(z_score) > threshold:
                anomalies.append((excel_row, excel_col, current_val, mean, std_dev, z_score,
                                  raw_year, raw_muni, raw_category, metric, hist_comment))

    # Build dictionaries (keyed by (municipality, category, metric)) for highlighting.
    # Build dictionarie for anomalies.
    anomaly_dict = {}
    for anomaly in anomalies:
        _, _, current_val, mean, std_dev, z_score, year, muni, category, metric, hist_comment = anomaly
        key = (str(muni).strip(), str(category).strip(), metric)
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        anomaly_dict[key] = (
            f"Anomaly Detected:\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nValue: {current_val:.2f}\nMean: {mean:.2f}\n"
            f"Std Dev: {std_dev:.2f}\nZ-Score: {z_score:.2f}\nHistorical Data:\n{hist_comment}"
        )
        
    # Build dictionaries for missing data.
    missing_dict = {}
    for m in missing_data:
        _, _, year, muni, category, metric, hist_comment = m
        key = (str(muni).strip(), str(category).strip(), metric)
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        missing_dict[key] = (
            f"Missing Data:\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nHistorical Data:\n{hist_comment}"
        )
        
    # Build dictionarie for non-numeric data.
    non_numeric_dict = {}
    for n in non_numeric:
        _, _, value, year, muni, category, metric = n
        key = (str(muni).strip(), str(category).strip(), metric)
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        non_numeric_dict[key] = (
            f"Non-Numeric Value:\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nValue: {value}"
        )

    # Transform the data into a pivot table using only rows with a category.
    pivot_df = transform_data(df_detection, category_col)

    # Write sheets to Excel.
    transformed_sheet_names = []
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write transformed data; if split_by is specified, split accordingly.
        if split_by is None or split_by.lower() not in ["municipality", "category"]:
            pivot_df.to_excel(writer, index=False, sheet_name="Transformed Data")
            transformed_sheet_names.append("Transformed Data")
        else:
            group_col = "Municipality" if split_by.lower() == "municipality" else "Category"
            for group_name, group_df in pivot_df.groupby(group_col):
                sheet_name = str(group_name)[:31]  # Ensure sheet name ≤ 31 characters
                group_df.to_excel(writer, index=False, sheet_name=sheet_name)
                transformed_sheet_names.append(sheet_name)
        # Write the original raw data (all rows) to a sheet called "Data"
        df.to_excel(writer, index=False, sheet_name="Data") 

    # Post-process with openpyxl: apply highlighting and create summary.
    wb = load_workbook(output_path)
    ws_data = wb["Data"]

    # Put back the second header row with actual metric names in the "Data" sheet.
    ws_data.insert_rows(2)
    for col in range(1, ws_data.max_column + 1):
        header_val = ws_data.cell(row=1, column=col).value
        if col > 3 and header_val in metric_names:
            ws_data.cell(row=2, column=col).value = metric_names.get(header_val, "")
        else:
            ws_data.cell(row=2, column=col).value = ""

    # Apply highlighting on the "Data" sheet.
    for anomaly in anomalies:
        excel_row, excel_col, current_val, mean, std_dev, z_score, year, muni, category, metric, hist_comment = anomaly
        cell = ws_data.cell(row=excel_row + 1, column=excel_col)
        cell.fill = red
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        cell.comment = Comment(
            f"Anomaly:\nMunicipality: {muni}\nCategory: {category}\nMetric: {metric_note}\n"
            f"Value: {current_val:.2f}\nMean: {mean:.2f}\nStd Dev: {std_dev:.2f}\n"
            f"Z-Score: {z_score:.2f}\nHistorical Data:\n{hist_comment}",
            "AI Detection Algorithm", width=350, height=300
        )
    for m in missing_data:
        excel_row, excel_col, year, muni, category, metric, hist_comment = m
        cell = ws_data.cell(row=excel_row + 1, column=excel_col)
        cell.fill = blue
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        cell.comment = Comment(
            f"Missing Data:\nMunicipality: {muni}\nCategory: {category}\nMetric: {metric_note}\n"
            f"Historical Data:\n{hist_comment}",
            "AI Detection Algorithm", width=300, height=200
        )
    for n in non_numeric:
        excel_row, excel_col, value, year, muni, category, metric = n
        cell = ws_data.cell(row=excel_row + 1, column=excel_col)
        cell.fill = yellow
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        cell.comment = Comment(
            f"Non-Numeric Value:\nMunicipality: {muni}\nCategory: {category}\nMetric: {metric_note}\n"
            f"Value: {value}",
            "AI Detection Algorithm", width=300, height=100
        )
    
    # Mark valid (non-flagged) cells in "Data" as green.
    flagged = set()
    for lst in (anomalies, missing_data, non_numeric):
        for entry in lst:
            flagged.add((entry[0] + 1, entry[1]))  # Adjust row by +1 for the inserted header row.
    
    # For each row in the raw data, only mark as green if the row has a category.
    for idx, row in df.iterrows():
        try:
            current_year = float(row[year_col])
        except (ValueError, TypeError):
            continue
        if current_year != target_year:
            continue
        # Skip rows with missing category.
        if pd.isna(row[category_col]):
            continue
        for metric in metric_cols:
            excel_row = idx + 3  # idx+2 then +1 for inserted header row.
            excel_col = df.columns.get_loc(metric) + 1
            if (excel_row, excel_col) in flagged:
                continue
            val = row[metric]
            if pd.isna(val):
                continue
            try:
                float(val)
                ws_data.cell(row=excel_row, column=excel_col).fill = green
            except (ValueError, TypeError):
                continue

    # Apply highlighting to each transformed data sheet.
    for sheet_name in transformed_sheet_names:
        ws_trans = wb[sheet_name]
        header_row = next(ws_trans.iter_rows(min_row=1, max_row=1))
        col_target = None
        for cell in header_row:
            if str(cell.value).strip() == str(target_year):
                col_target = cell.column
                break
        if col_target is not None:
            for row in ws_trans.iter_rows(min_row=2, values_only=False):
                muni_val = str(row[0].value).strip() if row[0].value is not None else "" # If it is null, it will be an empty string
                category_val = str(row[1].value).strip() if row[1].value is not None else ""
                metric_val = str(row[2].value).strip() if row[2].value is not None else ""
                key = (muni_val, category_val, metric_val)
                cell = row[col_target - 1]  # adjust for 0-indexing
                if key in anomaly_dict:
                    cell.fill = red
                    cell.comment = Comment(anomaly_dict[key], "AI Detection Algorithm", width=350, height=300)
                elif key in missing_dict:
                    cell.fill = blue
                    cell.comment = Comment(missing_dict[key], "AI Detection Algorithm", width=300, height=200)
                elif key in non_numeric_dict:
                    cell.fill = yellow
                    cell.comment = Comment(non_numeric_dict[key], "AI Detection Algorithm", width=300, height=100)
                else:
                    try:
                        if cell.value is not None:
                            float(cell.value)
                            cell.fill = green
                    except (ValueError, TypeError):
                        pass

    # Creates an "Events Summary" sheet that combines all data points flagged by algorithm.
    ws_summary = wb.create_sheet("Events Summary")
    summary_headers = ["Municipality", "Category", "Metric", "Type", "Value"]
    for col_idx, header in enumerate(summary_headers, start=1):
        ws_summary.cell(row=1, column=col_idx, value=header)
    summary_data = []
    
    # Append anomaly data points.
    for anomaly in anomalies:
        _, _, current_val, mean, std_dev, z_score, year, muni, category, metric, hist_comment = anomaly
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        note_text = (
            f"Anomaly Detected:\nYear: {year}\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nValue: {current_val:.2f}\nMean: {mean:.2f}\n"
            f"Std Dev: {std_dev:.2f}\nZ-Score: {z_score:.2f}\nHistorical Data:\n{hist_comment}"
        )
        summary_data.append((muni, category, metric_note, "Anomaly", current_val, note_text))
        
    # Append missing data data points.
    for m in missing_data:
        _, _, year, muni, category, metric, hist_comment = m
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        note_text = (
            f"Missing Data:\nYear: {year}\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nHistorical Data:\n{hist_comment}"
        )
        summary_data.append((muni, category, metric_note, "Missing", "N/A", note_text))
        
    # Append non-numeric data points.
    for n in non_numeric:
        _, _, value, year, muni, category, metric = n
        metric_note = f"{metric} - {metric_names.get(metric, '')}"
        note_text = (
            f"Non-Numeric Value:\nYear: {year}\nMunicipality: {muni}\nCategory: {category}\n"
            f"Metric: {metric_note}\nValue: {value}"
        )
        summary_data.append((muni, category, metric_note, "Non-Numeric", value, note_text))
    
    # Sort alphabetically by first element
    summary_data.sort(key=lambda x: x[0])
    for row_idx, row_data in enumerate(summary_data, start=2):
        ws_summary.cell(row=row_idx, column=1, value=row_data[0])
        ws_summary.cell(row=row_idx, column=2, value=row_data[1])
        ws_summary.cell(row=row_idx, column=3, value=row_data[2])
        ws_summary.cell(row=row_idx, column=4, value=row_data[3])
        value_cell = ws_summary.cell(row=row_idx, column=5, value=row_data[4])
        event_type = row_data[3]
        if event_type == "Anomaly":
            value_cell.fill = red
            value_cell.comment = Comment(row_data[5], "Algorithm", width=350, height=300)
        elif event_type == "Missing":
            value_cell.fill = blue
            value_cell.comment = Comment(row_data[5], "Algorithm", width=300, height=200)
        elif event_type == "Non-Numeric":
            value_cell.fill = yellow
            value_cell.comment = Comment(row_data[5], "Algorithm", width=300, height=100)

    wb.save(output_path)
    print(f"Anomaly detection and transformation complete. Output saved to: {output_path}")

def transform_data(df, category_col):
    """
    Pivot the raw data so that each row corresponds to a unique (Municipality [1st column], Category [2nd collumn], Metric [3rd column])
    and columns > 3 represent years. 
    """
    transformed = []
    for _, row in df.iterrows():
        municipality = row['a_jurisdiction']
        year = row['a_fiscal_year']
        for metric in df.columns[3:]:
            category = row[category_col]
            value = row[metric]
            transformed.append([municipality, category, metric, year, value])
            
    # Creates a DataFrame from the transformed data
    transformed_df = pd.DataFrame(transformed, columns=['Municipality', 'Category', 'Metric', 'Year', 'Value'])
    
    # Pivots the DataFrame, switches the years to columns
    pivot_df = transformed_df.pivot_table(
        index=['Municipality', 'Category', 'Metric'],
        columns='Year',
        values='Value',
        aggfunc='first'
    )
    
    pivot_df.reset_index(inplace=True)
    pivot_df.columns.name = None
    pivot_df.columns = ['Municipality', 'Category', 'Metric'] + [str(col) for col in pivot_df.columns[3:]]
    return pivot_df
